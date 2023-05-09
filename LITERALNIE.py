import random
from openpyxl import load_workbook

class backend:
#wygenerowanie losowej liczby porządkowej i słowa z pliku testowego excel
    def Generate_word():
        l_porzadkowa = random.randint(1,5)
        
        # Podczyt słowa 5 literowego z excela
        wb = load_workbook('SLOWA 5 LITEROWE.xlsx')
        sheet = wb.active
        
        szukane_słowo = sheet.cell(row = l_porzadkowa, column = 2).value
        print('Word generating successful')

        #stworzenie listy z szukanym słowem
        lista_szukane_słowo = []
        
        for x in szukane_słowo:
            lista_szukane_słowo.append(x)
        return lista_szukane_słowo

##      ***GAME LOGIC***
#zabezpieczenie przed wpisaniem złego słowa
    def User_Input():
        while True:             
            guess = input('Type your guess: ')
            if len(guess) == 5:
                print("Word lenght correct")
                break    
            print("Wrong word lenght, input 5 letter word.")

#stworzenie listy ze zgadywanym słowem 1
        lista_guess=[]
        for x in guess:
            x = x.upper()
            lista_guess.append(x)

        return lista_guess
    
#porównywanie liter w obydwu słowach    
    def Logika_gry():
        szukane_słowo_hint = ['_','_','_','_','_']
        lista_szukane_słowo = backend.Generate_word()
        
        while szukane_słowo_hint != lista_szukane_słowo:
            lista_guess = backend.User_Input()
            i=0
            while i<=4:
                for x in lista_guess:
                    if x == lista_szukane_słowo[i]:
                        szukane_słowo_hint[i] = x
                i = i+1
            print('Poszukiwane słowo jest postaci: ', szukane_słowo_hint)

        output = ''
        for letter in szukane_słowo_hint:
            output += letter
        print('Congratulations! Your word was: ', output)
        return szukane_słowo_hint
#####
Game_rules = ' 1. Generate random polish 5 letter word and keep it hidden \n 2. Give a user 6 chances to guess the word \n 3. Everytime a user takes a guess tell him if any letter in a current guess is included in the destination word if so, tell if it is in a correct position or wrong position'
print(Game_rules)
backend.Logika_gry()
